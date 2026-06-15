# /// script
# requires-python = ">=3.12"
# dependencies = [
#   "numpy",
#   "pandas",
#   "polars",
#   "pyarrow",
#   "openpyxl",
#   "matplotlib",
# ]
# ///
"""Storage benchmark for the Storage chapter.

Answers one realistic question — a filtered group-by aggregation —
    SELECT category, COUNT(*), AVG(value), SUM(amount)
    FROM data WHERE amount > 100 GROUP BY category
against the same synthetic data held five different ways, across a sweep of
row counts, and plots time-vs-rows on a log-log chart.

Methodology (kept out of the chapter on purpose — the chart is the only trace):
  * File formats (CSV, Parquet, Excel) are RE-READ on every query: that is the
    reality of a pile of files. pandas reads eagerly; polars scans lazily
    (projection + predicate pushdown), which is each library's idiomatic use.
  * SQLite is loaded ONCE (untimed) and then queried — which is the entire
    point of using a store. It gets a covering index on the columns the query
    touches, so the row store is not handicapped against columnar Parquet.
  * Excel is marked DNF beyond its hard 1,048,576-row limit.

Run:  uv run benchmark.py     (first run pulls wheels)
"""

from __future__ import annotations

import json
import multiprocessing as mp
import os
import sqlite3
import statistics
import time
from pathlib import Path

import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
# Override at smaller scale with e.g. BENCH_ROWS=1000,10000 for a smoke test.
ROW_COUNTS = [int(x) for x in
              os.environ.get("BENCH_ROWS", "1000,10000,100000,1000000,10000000").split(",")]
EXCEL_ROW_LIMIT = 1_048_576  # Excel's hard cap; beyond this we mark DNF
TIMEOUT_S = 360  # per-measurement wall-clock guard
SEED = 42

HERE = Path(__file__).resolve().parent
WORK = HERE / "_work"
OUT_PNG = HERE.parents[1] / "course" / "docs" / "assets" / "storage-benchmark.png"
OUT_JSON = HERE / "results.json"

# The query, expressed once for SQLite.
SQL = ("SELECT category, COUNT(*) AS n, AVG(value) AS av, SUM(amount) AS sm "
       "FROM data WHERE amount > 100 GROUP BY category")

CONTENDERS = [
    "CSV + polars",
    "Parquet + polars",
    "Excel (xlsx)",
    "SQLite",
]


# ---------------------------------------------------------------------------
# Data + artifacts
# ---------------------------------------------------------------------------
def generate(n: int) -> pd.DataFrame:
    rng = np.random.default_rng(SEED)
    cats = np.array([f"cat_{i:02d}" for i in range(50)])
    regions = np.array([f"region_{i}" for i in range(8)])
    return pd.DataFrame({
        "id": np.arange(n, dtype=np.int64),
        "category": cats[rng.integers(0, 50, n)],
        "region": regions[rng.integers(0, 8, n)],
        "value": rng.random(n),
        "amount": rng.exponential(100.0, n),
        "score": rng.normal(0.0, 1.0, n),
        "flag": rng.integers(0, 2, n),
    })


def paths() -> dict[str, Path]:
    return {
        "csv": WORK / "data.csv",
        "parquet": WORK / "data.parquet",
        "xlsx": WORK / "data.xlsx",
        "sqlite": WORK / "data.sqlite",
    }


def write_xlsx(df: pd.DataFrame, path: Path) -> None:
    """Fast, low-memory xlsx write via openpyxl's write_only mode."""
    from openpyxl import Workbook

    wb = Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    wb.save(path)


def build_artifacts(df: pd.DataFrame, p: dict[str, Path], do_excel: bool) -> None:
    df.to_csv(p["csv"], index=False)
    df.to_parquet(p["parquet"], index=False)

    if p["sqlite"].exists():
        p["sqlite"].unlink()
    con = sqlite3.connect(p["sqlite"])
    df.to_sql("data", con, index=False, if_exists="replace")
    # Covering index on exactly the columns the query touches: lets SQLite scan
    # a narrow, category-ordered B-tree (no table-row lookups, no GROUP BY sort)
    # — the row-store analogue of the column-skipping columnar Parquet gets free.
    con.execute("CREATE INDEX idx_cover ON data(category, amount, value)")
    con.execute("ANALYZE")
    con.close()

    if do_excel:
        write_xlsx(df, p["xlsx"])


# ---------------------------------------------------------------------------
# The operation, per contender (each returns the aggregated result, discarded)
# ---------------------------------------------------------------------------
def op_csv_polars(p):
    import polars as pl

    return (pl.scan_csv(p["csv"]).filter(pl.col("amount") > 100)
            .group_by("category")
            .agg(pl.len().alias("n"), pl.col("value").mean(), pl.col("amount").sum())
            .collect())


def op_parquet_polars(p):
    import polars as pl

    return (pl.scan_parquet(p["parquet"]).filter(pl.col("amount") > 100)
            .group_by("category")
            .agg(pl.len().alias("n"), pl.col("value").mean(), pl.col("amount").sum())
            .collect())


def op_excel(p):
    from openpyxl import load_workbook

    wb = load_workbook(p["xlsx"], read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    df = pd.DataFrame(rows, columns=header)
    wb.close()
    d = df[df["amount"] > 100]
    return d.groupby("category").agg(n=("id", "size"), av=("value", "mean"), sm=("amount", "sum"))


def op_sqlite(p):
    con = sqlite3.connect(p["sqlite"])
    out = con.execute(SQL).fetchall()
    con.close()
    return out


OPS = {
    "CSV + polars": op_csv_polars,
    "Parquet + polars": op_parquet_polars,
    "Excel (xlsx)": op_excel,
    "SQLite": op_sqlite,
}


def runs_for(contender: str, n: int) -> int:
    if contender == "Excel (xlsx)":
        return 1 if n >= 1_000_000 else (2 if n >= 100_000 else 3)
    if n >= 10_000_000:
        return 2
    if n >= 1_000_000:
        return 3
    return 5


# ---------------------------------------------------------------------------
# Timeout-guarded measurement (subprocess so a hang can be killed)
# ---------------------------------------------------------------------------
def _worker(q, contender, p, k):
    try:
        op = OPS[contender]
        op(p)  # warm-up (cache, JIT)
        times = []
        for _ in range(k):
            t0 = time.perf_counter()
            op(p)
            times.append(time.perf_counter() - t0)
        q.put(("ok", statistics.median(times)))
    except Exception as exc:  # noqa: BLE001
        q.put(("err", repr(exc)))


def measure(contender: str, p: dict, n: int) -> float | None:
    ctx = mp.get_context("spawn")
    q = ctx.Queue()
    proc = ctx.Process(target=_worker, args=(q, contender, p, runs_for(contender, n)))
    proc.start()
    proc.join(TIMEOUT_S)
    if proc.is_alive():
        proc.terminate()
        proc.join()
        print(f"      {contender}: TIMEOUT (>{TIMEOUT_S}s) -> DNF")
        return None
    status, payload = q.get()
    if status == "err":
        print(f"      {contender}: ERROR {payload} -> DNF")
        return None
    print(f"      {contender}: {payload * 1000:.1f} ms")
    return payload


# ---------------------------------------------------------------------------
# Plot
# ---------------------------------------------------------------------------
def plot(results: dict[str, dict[int, float | None]]) -> None:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(9.5, 6.2))
    markers = ["s", "^", "v", "D"]
    cmap = plt.get_cmap("tab10")

    for i, contender in enumerate(CONTENDERS):
        series = results[contender]
        xs = [n for n in ROW_COUNTS if series.get(n) is not None]
        ys = [series[n] for n in xs]
        if not xs:
            continue
        ax.plot(xs, ys, marker=markers[i], color=cmap(i), label=contender,
                linewidth=2, markersize=7)

    # Mark Excel DNF where it ran out of rows.
    excel_dnf = [n for n in ROW_COUNTS if n > EXCEL_ROW_LIMIT]
    if excel_dnf:
        ax.annotate("Excel: DNF\n(> 1,048,576 rows)",
                    xy=(excel_dnf[0], ax.get_ylim()[0]),
                    xytext=(excel_dnf[0], ax.get_ylim()[0] * 3),
                    ha="center", fontsize=8, color=cmap(2))

    ax.set_xscale("log")
    ax.set_yscale("log")
    ax.set_xlabel("rows")
    ax.set_ylabel("time per query (seconds, median)")
    ax.set_title("Time to answer one aggregate query vs. dataset size\n"
                 "filtered GROUP BY — files re-read each query, SQLite pre-loaded & indexed — lower is better",
                 fontsize=11)
    ax.grid(True, which="both", linestyle=":", alpha=0.5)
    ax.legend(fontsize=9)
    fig.tight_layout()
    OUT_PNG.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(OUT_PNG, dpi=150)
    print(f"\nchart -> {OUT_PNG}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    WORK.mkdir(parents=True, exist_ok=True)
    p = paths()
    results: dict[str, dict[int, float | None]] = {c: {} for c in CONTENDERS}

    for n in ROW_COUNTS:
        print(f"\n=== {n:,} rows ===")
        do_excel = n <= EXCEL_ROW_LIMIT
        print("  building artifacts...")
        df = generate(n)
        build_artifacts(df, p, do_excel)
        del df

        for contender in CONTENDERS:
            if contender == "Excel (xlsx)" and not do_excel:
                print(f"      {contender}: DNF (over row limit)")
                results[contender][n] = None
                continue
            results[contender][n] = measure(contender, p, n)

    OUT_JSON.write_text(json.dumps(
        {c: {str(k): v for k, v in s.items()} for c, s in results.items()}, indent=2))
    print(f"\nresults -> {OUT_JSON}")
    plot(results)


if __name__ == "__main__":
    main()
